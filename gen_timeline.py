import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import date, timedelta, datetime
import calendar

# --- Configuration ---
# Note: Adjust these based on the actual image content if needed
start_year = 2025 # Or the year from your image
start_month = 5   # Or the starting month from your image
num_months = 24    # How many months to display horizontally

output_filename = "project_schedule.xlsx"
task_col_count = 3 # Number of columns reserved for task names/info
header_row_count = 3 # Number of rows for Month/Day/Weekday headers

# Define tasks (Example data - replace with your actual tasks)
# Format: {'name': 'Task Name', 'row_offset': row_index_below_headers,
#          'start': 'YYYY-MM-DD', 'end': 'YYYY-MM-DD', 'color': 'HEXCOLOR'}
tasks = [
    {'name': '基本計画 (Basic Planning)', 'row_offset': 1, 'start': '2025-04-03', 'end': '2025-04-15', 'color': 'FFFF00'}, # Yellow
    {'name': '基本設計 (Basic Design)', 'row_offset': 2, 'start': '2025-04-16', 'end': '2025-05-10', 'color': 'FFC0CB'}, # Pink
    {'name': '- プラン提案 (Plan Proposal)', 'row_offset': 3, 'start': '2025-04-16', 'end': '2025-04-30', 'color': 'ADD8E6'}, # Light Blue
    {'name': '- デザイン検討 (Design Review)', 'row_offset': 4, 'start': '2025-05-01', 'end': '2025-05-10', 'color': 'ADD8E6'}, # Light Blue
    {'name': '実施設計 (Detailed Design)', 'row_offset': 5, 'start': '2025-05-11', 'end': '2025-06-20', 'color': 'FF8C00'}, # Dark Orange
    {'name': '- 詳細図面作成 (Detail Drawings)', 'row_offset': 6, 'start': '2025-05-11', 'end': '2025-06-10', 'color': '90EE90'}, # Light Green
    {'name': '- 仕様確認 (Spec Check)', 'row_offset': 7, 'start': '2025-06-11', 'end': '2025-06-20', 'color': '90EE90'}, # Light Green
]

# --- Helper Functions & Setup ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Schedule"

# Styles
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
header_font = Font(bold=True)

# Japanese weekdays
weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]

# --- Generate Headers and Date Grid ---
current_col = task_col_count + 1
date_to_col = {} # Dictionary to map date objects to column numbers

print("Generating headers...")
for i in range(num_months):
    year = start_year + (start_month + i - 1) // 12
    month = (start_month + i - 1) % 12 + 1
    month_name = f"{year} / {month}月"

    # Get number of days in the month
    days_in_month = calendar.monthrange(year, month)[1]

    # Write Month header and merge
    month_start_col = current_col
    month_end_col = month_start_col + days_in_month - 1
    ws.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=month_end_col)
    month_cell = ws.cell(row=1, column=month_start_col, value=month_name)
    month_cell.alignment = center_align
    month_cell.font = header_font
    month_cell.border = thin_border


    # Write Day and Weekday headers
    for day_num in range(1, days_in_month + 1):
        current_date = date(year, month, day_num)
        col_idx = month_start_col + day_num - 1

        # Day number
        day_cell = ws.cell(row=2, column=col_idx, value=day_num)
        day_cell.alignment = center_align
        day_cell.border = thin_border

        # Weekday
        weekday_str = weekdays_jp[current_date.weekday()] # Monday is 0
        weekday_cell = ws.cell(row=3, column=col_idx, value=weekday_str)
        weekday_cell.alignment = center_align
        weekday_cell.border = thin_border

        # Apply weekend shading (optional)
        if current_date.weekday() == 5: # Saturday
             day_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
             weekday_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        elif current_date.weekday() == 6: # Sunday
             day_cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
             weekday_cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")


        date_to_col[current_date] = col_idx # Store mapping

    current_col = month_end_col + 1

print("Date to column mapping complete.")
# Set column widths (adjust as needed)
ws.column_dimensions['A'].width = 25 # Wider column for task names
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
for col_letter in [openpyxl.utils.get_column_letter(i) for i in range(task_col_count + 1, current_col)]:
    ws.column_dimensions[col_letter].width = 3.5 # Narrow columns for dates


# --- Add Task Data and Timeline Bars ---
max_task_row = header_row_count + len(tasks) # Determine grid height dynamically if needed
print(f"Adding {len(tasks)} tasks...")

for task in tasks:
    task_row_index = header_row_count + task['row_offset']

    # Write task name (assuming it goes in the first column)
    task_name_cell = ws.cell(row=task_row_index, column=1, value=task['name'])
    task_name_cell.alignment = left_align
    task_name_cell.border = thin_border # Apply border to task name cell too

    # Find start and end columns for the timeline bar
    try:
        start_dt = datetime.strptime(task['start'], '%Y-%m-%d').date()
        end_dt = datetime.strptime(task['end'], '%Y-%m-%d').date()
    except ValueError:
        print(f"Warning: Could not parse dates for task '{task['name']}'. Skipping timeline bar.")
        continue

    start_col = date_to_col.get(start_dt)
    end_col = date_to_col.get(end_dt)

    if start_col is not None and end_col is not None and start_col <= end_col :
        # Merge cells for the timeline bar
        print(f"  - Adding bar for '{task['name']}' from {start_dt} (col {start_col}) to {end_dt} (col {end_col}) in row {task_row_index}")
        ws.merge_cells(start_row=task_row_index, start_column=start_col,
                       end_row=task_row_index, end_column=end_col)

        # Apply color and border to the merged range (apply to top-left cell)
        merged_cell = ws.cell(row=task_row_index, column=start_col)
        merged_cell.fill = PatternFill(start_color=task['color'], end_color=task['color'], fill_type="solid")
        merged_cell.border = thin_border
        # Add alignment if you want text inside the bar
        merged_cell.alignment = center_align
        # Optionally add text to the bar: merged_cell.value = "Label"

        # Need to apply border formatting to *all* cells within the merged range
        # if we want the border to show correctly when cells are unmerged later
        # or for better visual consistency in some viewers.
        for r in range(task_row_index, task_row_index + 1):
             for c in range(start_col, end_col + 1):
                  cell_in_range = ws.cell(row=r, column=c)
                  cell_in_range.border = thin_border
                  # Ensure fill color is applied too, in case merge doesn't propagate it perfectly everywhere
                  # cell_in_range.fill = PatternFill(start_color=task['color'], end_color=task['color'], fill_type="solid")

    else:
        print(f"Warning: Could not find columns for task '{task['name']}' dates ({start_dt} to {end_dt}). Dates might be out of generated range.")

# Apply borders to the rest of the grid (optional, for visual completeness)
print("Applying grid borders...")
first_date_col = task_col_count + 1
last_date_col = current_col -1
for r in range(header_row_count + 1, max_task_row + 1):
    # Task info columns border (if not already done)
    for c in range(1, task_col_count + 1):
         cell = ws.cell(row=r, column=c)
         if not cell.border or cell.border == thin_border: # Avoid overriding existing borders unnecessarily
              cell.border = thin_border
    # Date grid area borders
    for c in range(first_date_col, last_date_col + 1):
        cell = ws.cell(row=r, column=c)
        # Only apply border if it's not part of a filled timeline bar
        # (or apply thin_border regardless if you want grid lines over bars too)
        if (cell.fill is None or cell.fill.fill_type is None or cell.fill.fill_type == 'none'):
            cell.border = thin_border


# --- Save the Workbook ---
try:
    wb.save(output_filename)
    print(f"Excel file '{output_filename}' generated successfully.")
except PermissionError:
    print(f"Error: Permission denied. Make sure '{output_filename}' is not already open in Excel.")
except Exception as e:
    print(f"An error occurred while saving the file: {e}")